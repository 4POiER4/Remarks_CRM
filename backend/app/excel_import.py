from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.models import Letter, ProjectObject, Remark

COLUMN_MAP = {
  "от кого письмо/задание": "from_whom",
  "от кого письмо": "from_whom",
  "от кого": "from_whom",
  "№ письма": "letter_number",
  "номер письма": "letter_number",
  "n письма": "letter_number",
  "дата письма": "letter_date",
  "сопровод лэп": "lep_accompaniment",
  "сопровод леп": "lep_accompaniment",
  "дата сопровода лэп": "lep_accompaniment_date",
  "дата сопровода леп": "lep_accompaniment_date",
  "объект": "object_name",
  "основной объект": "object_name",
  "подобъект": "subobject_name",
  "подобъект/объект": "subobject_name",
  "замечание к документу": "document_remark",
  "вид документа": "document_type",
  "замечание": "remark_text",
  "от кого письмо/задание": "from_whom",
  "от кого письмо": "from_whom",
  "№ письма": "letter_number",
  "номер письма": "letter_number",
  "дата письма": "letter_date",
  "сопровод лэп": "lep_accompaniment",
  "дата сопровода лэп": "lep_accompaniment_date",
  "объект": "object_name",
  "основной объект": "object_name",
  "подобъект": "subobject_name",
  "подобъект/объект": "subobject_name",
  "замечание к документу": "document_remark",
  "вид документа": "document_type",
  "замечание": "remark_text",
}


def _normalize_header(value) -> str:
  if value is None:
    return ""
  return " ".join(str(value).replace("\xa0", " ").strip().lower().replace("\u0451", "\u0435").split())


def _parse_date(value) -> date | None:
  if value is None or value == "":
    return None
  if isinstance(value, datetime):
    return value.date()
  if isinstance(value, date):
    return value
  text = str(value).strip()
  for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
    try:
      return datetime.strptime(text, fmt).date()
    except ValueError:
      continue
  return None


def _parse_text(value) -> str | None:
  if value is None:
    return None
  text = str(value).strip()
  return text or None


def _get_or_create_object(
  db: Session,
  cache: dict[tuple[str, str | None], ProjectObject],
  name: str,
  subobject_name: str | None,
) -> ProjectObject:
  key = (name, subobject_name)
  if key not in cache:
    obj = (
      db.query(ProjectObject)
      .filter(ProjectObject.name == name, ProjectObject.subobject_name == subobject_name)
      .first()
    )
    if not obj:
      obj = ProjectObject(name=name, subobject_name=subobject_name)
      db.add(obj)
      db.flush()
    cache[key] = obj
  return cache[key]


def _get_or_create_letter(db: Session, cache: dict[tuple, Letter], obj: ProjectObject, data: dict) -> Letter:
  key = (
    obj.id,
    data.get("from_whom") or "",
    data.get("letter_number") or "",
    str(data.get("letter_date") or ""),
    data.get("lep_accompaniment") or "",
    str(data.get("lep_accompaniment_date") or ""),
  )
  if key not in cache:
    letter = (
      db.query(Letter)
      .filter(
        Letter.object_id == obj.id,
        Letter.from_whom == data.get("from_whom"),
        Letter.letter_number == data.get("letter_number"),
        Letter.letter_date == data.get("letter_date"),
        Letter.lep_accompaniment == data.get("lep_accompaniment"),
        Letter.lep_accompaniment_date == data.get("lep_accompaniment_date"),
      )
      .first()
    )
    if not letter:
      letter = Letter(
        object_id=obj.id,
        from_whom=data.get("from_whom"),
        letter_number=data.get("letter_number"),
        letter_date=data.get("letter_date"),
        lep_accompaniment=data.get("lep_accompaniment"),
        lep_accompaniment_date=data.get("lep_accompaniment_date"),
      )
      db.add(letter)
      db.flush()
    cache[key] = letter
  return cache[key]


def import_remarks_from_excel(content: bytes, db: Session) -> tuple[int, int, list[str]]:
  workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
  sheet = workbook.active
  rows = sheet.iter_rows(values_only=True)

  field_indexes: dict[str, int] = {}
  header_row_number = 0
  for row_number, header_row in enumerate(rows, start=1):
    candidate_indexes: dict[str, int] = {}
    for index, cell in enumerate(header_row):
      normalized = _normalize_header(cell)
      if normalized in COLUMN_MAP:
        candidate_indexes[COLUMN_MAP[normalized]] = index
    if len(candidate_indexes) >= 2 or {"document_remark", "remark_text"} & set(candidate_indexes):
      field_indexes = candidate_indexes
      header_row_number = row_number
      break
    if row_number >= 30:
      break

  if not field_indexes:
    return 0, 0, ["Не найдены заголовки столбцов. Проверьте первую строку Excel."]

  imported = 0
  skipped = 0
  errors: list[str] = []
  object_cache: dict[tuple[str, str | None], ProjectObject] = {}
  letter_cache: dict[tuple, Letter] = {}

  for row_number, row in enumerate(rows, start=header_row_number + 1):
    if not row or all(cell is None or str(cell).strip() == "" for cell in row):
      skipped += 1
      continue

    data: dict[str, object] = {}
    for field, index in field_indexes.items():
      if index >= len(row):
        continue
      value = row[index]
      if field in {"letter_date", "lep_accompaniment_date"}:
        data[field] = _parse_date(value)
      else:
        data[field] = _parse_text(value)

    if not any(data.values()):
      skipped += 1
      continue

    try:
      object_name = str(data.pop("object_name", None) or "").strip() or "Без объекта"
      subobject_name = str(data.pop("subobject_name", None) or "").strip() or None

      obj = _get_or_create_object(db, object_cache, object_name, subobject_name)
      letter_data = {
        "from_whom": data.get("from_whom"),
        "letter_number": data.get("letter_number"),
        "letter_date": data.get("letter_date"),
        "lep_accompaniment": data.get("lep_accompaniment"),
        "lep_accompaniment_date": data.get("lep_accompaniment_date"),
      }
      letter = _get_or_create_letter(db, letter_cache, obj, letter_data)
      remark = Remark(
        letter_id=letter.id,
        document_remark=data.get("document_remark"),
        document_type=data.get("document_type"),
        remark_text=data.get("remark_text"),
      )
      db.add(remark)
      imported += 1
      if imported % 200 == 0:
        db.commit()
    except Exception as exc:
      errors.append(f"Строка {row_number}: {exc}")

  if imported:
    db.commit()

  return imported, skipped, errors
