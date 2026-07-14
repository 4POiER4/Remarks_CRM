from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import Remark

COLUMN_MAP = {
    "от кого письмо/задание": "from_whom",
    "от кого письмо": "from_whom",
    "№ письма": "letter_number",
    "номер письма": "letter_number",
    "дата письма": "letter_date",
    "сопровод лэп": "lep_accompaniment",
    "дата сопровода лэп": "lep_accompaniment_date",
    "объект": "object_name",
    "замечание к документу": "document_remark",
    "вид документа": "document_type",
    "замечание": "remark_text",
}


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def import_remarks_from_excel(content: bytes, db: Session) -> tuple[int, int, list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        return 0, 0, ["Файл пуст"]

    field_indexes: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        if normalized in COLUMN_MAP:
            field_indexes[COLUMN_MAP[normalized]] = index

    if not field_indexes:
        return 0, 0, ["Не найдены заголовки столбцов. Проверьте первую строку Excel."]

    imported = 0
    skipped = 0
    errors: list[str] = []

    for row_number, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            skipped += 1
            continue

        data: dict[str, object] = {}
        for field, index in field_indexes.items():
            if index >= len(row):
                continue
            value = row[index]
            if field in {"letter_date", "lep_accompaniment_date"}:
                data[field] = _parse_date(value)
            else:
                data[field] = _parse_text(value)

        if not any(data.values()):
            skipped += 1
            continue

        try:
            remark = Remark(**data)
            db.add(remark)
            imported += 1
        except Exception as exc:
            errors.append(f"Строка {row_number}: {exc}")

    if imported:
        db.commit()

    return imported, skipped, errors
